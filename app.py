import io
import re
from datetime import datetime
from pathlib import Path

import matplotlib as mpl
import matplotlib.font_manager as fm
import matplotlib.pyplot as plt
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

st.set_page_config(page_title="Excel 汇总工具", page_icon="📊", layout="centered")

FIXED_LIST = [
    "罗鑫",
    "邓博文",
    "陈立",
    "李芝瑶",
    "张菁",
    "郑岑",
    "周钰萱",
    "丰亦舟",
    "吴筱烨",
    "范颖",
    "查泽民",
]

DEFAULT_SHEET_NAME = "Sheet1"
SOURCE_USECOLS = "G:P"  # G = name column, H:P = 9 metric columns
METRIC_COUNT = 9
DEFAULT_FILE_STEM = f"业绩表-{datetime.now().strftime('%Y%m%d')}"
DEFAULT_EXCEL_FILENAME = f"{DEFAULT_FILE_STEM}.xlsx"

APP_DIR = Path(__file__).parent
HOME_IMAGE_PATH = APP_DIR / "home.png"


@st.cache_resource
def setup_chinese_font() -> fm.FontProperties | None:
    """Locate a CJK-capable font, register it with matplotlib, and return it.

    Registering the font via ``fm.fontManager.addfont`` is what makes the
    subsequent ``rcParams`` assignment actually take effect; without
    registration matplotlib silently falls back to DejaVu Sans, which has no
    CJK glyphs and renders Chinese characters as empty boxes.

    Returns:
        FontProperties bound to the chosen font file, or None if no suitable
        font could be located.
    """
    # Bundled font files take priority — they ship with the repo so behaviour
    # is identical across local dev and Streamlit Community Cloud.
    candidate_paths = [
        APP_DIR / "SimHei.ttf",
        APP_DIR / "NotoSansCJKsc-Regular.otf",
        APP_DIR / "SourceHanSansCN-Regular.otf",
        APP_DIR / "fonts" / "SimHei.ttf",
        APP_DIR / "fonts" / "NotoSansCJKsc-Regular.otf",
        APP_DIR / "fonts" / "SourceHanSansCN-Regular.otf",
    ]

    chosen_path: Path | None = None
    for path in candidate_paths:
        if path.exists():
            chosen_path = path
            break

    # Fall back to a system-installed CJK font if nothing is bundled.
    if chosen_path is None:
        preferred_keywords = [
            "simhei",
            "notosanscjk",
            "noto sans cjk",
            "sourcehansans",
            "source han sans",
            "wenquanyi",
            "microsoft yahei",
            "pingfang",
            "arial unicode",
        ]
        system_fonts = fm.findSystemFonts(fontpaths=None, fontext="ttf") + fm.findSystemFonts(
            fontpaths=None, fontext="otf"
        )
        for font_path in system_fonts:
            lower_path = font_path.lower()
            if any(keyword in lower_path for keyword in preferred_keywords):
                chosen_path = Path(font_path)
                break

    if chosen_path is None:
        return None

    # Register so matplotlib's internal font lookup can find it by name.
    fm.fontManager.addfont(str(chosen_path))
    font_prop = fm.FontProperties(fname=str(chosen_path))
    font_name = font_prop.get_name()

    # Prepend our font so it wins over the default sans-serif stack.
    existing_stack = list(mpl.rcParams.get("font.sans-serif", []))
    if font_name in existing_stack:
        existing_stack.remove(font_name)
    mpl.rcParams["font.sans-serif"] = [font_name, *existing_stack]
    mpl.rcParams["font.family"] = "sans-serif"
    mpl.rcParams["axes.unicode_minus"] = False  # Render minus signs correctly.

    return font_prop


# Set up the font once at import time so every plot inherits the rcParams.
CHINESE_FONT_PROP = setup_chinese_font()


def sanitize_filename(filename: str, default_suffix: str) -> str:
    """Sanitize a user-supplied filename and ensure it ends with the given suffix.

    Args:
        filename: Raw filename from the user.
        default_suffix: Required extension, including the leading dot.

    Returns:
        A filesystem-safe filename guaranteed to end with ``default_suffix``.
    """
    name = (filename or "").strip()
    name = name.replace("\\", "_").replace("/", "_")
    name = re.sub(r'[<>:"|?*]', "_", name)
    if not name:
        name = f"业绩表-{datetime.now().strftime('%Y%m%d')}"

    path = Path(name)
    if path.suffix.lower() != default_suffix.lower():
        name = f"{path.stem or path.name}{default_suffix}"
    return name


def build_plot_filename(excel_filename: str) -> str:
    """Derive the PNG filename from a sanitized Excel filename."""
    sanitized_excel = sanitize_filename(excel_filename, ".xlsx")
    stem = Path(sanitized_excel).stem
    return f"{stem}.png"


def format_number(value) -> str:
    """Format a numeric value for display inside the rendered table image."""
    if pd.isna(value):
        return ""
    if isinstance(value, int):
        return f"{value:,}"
    if isinstance(value, float):
        if value.is_integer():
            return f"{int(value):,}"
        return f"{value:,.2f}".rstrip("0").rstrip(".")
    return str(value)


def dataframe_elementwise_map(df: pd.DataFrame, func) -> pd.DataFrame:
    """Apply ``func`` element-wise across ``df`` on both modern and legacy pandas.

    pandas >= 2.1 exposes ``DataFrame.map``; older versions only provide
    ``DataFrame.applymap``. This helper picks whichever is available.
    """
    if hasattr(df, "map"):
        return df.map(func)
    return df.applymap(func)


def process_excel(uploaded_file, sheet_name: str = DEFAULT_SHEET_NAME) -> tuple[bytes, pd.DataFrame]:
    """Aggregate the uploaded workbook against the fixed name list.

    Reads ``G:P`` from the chosen sheet, sums the 9 metric columns per name,
    aligns the result to ``FIXED_LIST`` order (filling absent names with 0),
    and writes a styled output workbook.

    Args:
        uploaded_file: File-like object yielded by ``st.file_uploader``.
        sheet_name: Worksheet to read from. Defaults to ``Sheet1``.

    Returns:
        Tuple of (xlsx bytes, result DataFrame) for downstream display.

    Raises:
        ValueError: If the sheet is missing or the column shape is unexpected.
    """
    uploaded_file.seek(0)

    try:
        df = pd.read_excel(
            uploaded_file,
            sheet_name=sheet_name,
            usecols=SOURCE_USECOLS,
            engine="openpyxl",
        )
    except ValueError as exc:
        raise ValueError(
            f"读取失败。请确认文件中存在工作表“{sheet_name}”，且 G:P 区域可读取。原始错误：{exc}"
        ) from exc

    expected_col_count = 10
    if df.shape[1] != expected_col_count:
        raise ValueError(f"读取到的列数为 {df.shape[1]}，但预期应为 10 列（G:P）。")

    headers = list(df.columns)
    name_col = headers[0]
    metric_cols = headers[1:]

    if len(metric_cols) != METRIC_COUNT:
        raise ValueError(f"数值指标列数量为 {len(metric_cols)}，但预期应为 {METRIC_COUNT}。")

    df[name_col] = df[name_col].fillna("").astype(str).str.strip()

    for col in metric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    df_filtered = df[df[name_col].isin(FIXED_LIST)].copy()

    if not df_filtered.empty:
        grouped = df_filtered.groupby(name_col, as_index=False)[metric_cols].sum()
    else:
        grouped = pd.DataFrame(columns=[name_col] + metric_cols)

    # Reindex against FIXED_LIST to preserve display order and zero-fill absentees.
    result_df = pd.DataFrame({name_col: FIXED_LIST})
    result_df = result_df.merge(grouped, on=name_col, how="left")
    for col in metric_cols:
        result_df[col] = result_df[col].fillna(0)

    # Cast back to int when every value in a column happens to be a whole number.
    for col in metric_cols:
        series = result_df[col]
        if (series % 1 == 0).all():
            result_df[col] = series.astype("int64")

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        result_df.to_excel(writer, index=False, sheet_name="处理结果")
        worksheet = writer.sheets["处理结果"]

        header_fill = PatternFill(fill_type="solid", fgColor="4F81BD")
        header_font = Font(bold=True, color="FFFFFF")
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_side = Side(style="thin", color="000000")
        thin_border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side,
        )
        zebra_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")

        max_row = worksheet.max_row
        max_col = worksheet.max_column

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border

        for row in worksheet.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.alignment = center_alignment
                cell.border = thin_border

        # Zebra striping: shade every other data row for readability.
        for row_idx in range(2, max_row + 1, 2):
            for col_idx in range(1, max_col + 1):
                worksheet.cell(row=row_idx, column=col_idx).fill = zebra_fill

        # Auto-size columns; CJK glyphs need extra width because they render wider.
        for col in worksheet.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                value = "" if cell.value is None else str(cell.value)
                display_width = len(value) + 2
                if any("\u4e00" <= char <= "\u9fff" for char in value):
                    display_width += 2
                max_length = max(max_length, display_width)
            worksheet.column_dimensions[column_letter].width = max_length

    output.seek(0)
    return output.getvalue(), result_df


def render_table_plot(result_df: pd.DataFrame) -> bytes:
    """Render the result DataFrame as a high-resolution PNG table.

    Each text element is explicitly bound to ``CHINESE_FONT_PROP`` so the CJK
    font is used even if matplotlib's family resolution decides otherwise.
    """
    font_prop = CHINESE_FONT_PROP

    display_df = result_df.copy()
    formatted_display_df = dataframe_elementwise_map(display_df, format_number)
    rows, cols = formatted_display_df.shape

    fig_width = max(12, cols * 1.8)
    fig_height = max(4.8, rows * 0.65 + 1.4)

    fig, ax = plt.subplots(figsize=(fig_width, fig_height), dpi=240)
    fig.patch.set_facecolor("white")
    ax.set_facecolor("white")
    ax.axis("off")

    table = ax.table(
        cellText=formatted_display_df.values,
        colLabels=list(formatted_display_df.columns),
        loc="center",
        cellLoc="center",
        colLoc="center",
    )

    table.auto_set_font_size(False)
    table.set_fontsize(12)
    table.scale(1.08, 1.55)

    for (row, _col), cell in table.get_celld().items():
        cell.set_edgecolor("#BFBFBF")
        cell.set_linewidth(0.8)
        cell.set_facecolor("white")
        txt = cell.get_text()
        txt.set_ha("center")
        txt.set_va("center")
        if font_prop is not None:
            txt.set_fontproperties(font_prop)
        if row == 0:
            cell.set_facecolor("#4F81BD")
            txt.set_color("white")
            txt.set_weight("bold")
        elif row % 2 == 0:
            cell.set_facecolor("#F8F8F8")

    plt.tight_layout(pad=0.8)

    output = io.BytesIO()
    fig.savefig(output, format="png", dpi=300, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    output.seek(0)
    return output.getvalue()


def reset_download_flags() -> None:
    """Clear download tracking when the user starts a fresh processing run."""
    st.session_state.excel_downloaded = False
    st.session_state.image_downloaded = False
    st.session_state.celebrated = False


def show_celebration() -> None:
    """Display the cute home.png with a cheering message after both downloads.

    The image is shown only once per processing run (``celebrated`` flag) so
    the balloons don't re-fire on every Streamlit rerun.
    """
    if not st.session_state.get("celebrated", False):
        st.balloons()
        st.session_state.celebrated = True

    st.markdown("---")
    if HOME_IMAGE_PATH.exists():
        col_left, col_center, col_right = st.columns([1, 2, 1])
        with col_center:
            st.image(str(HOME_IMAGE_PATH), use_container_width=True)
    else:
        st.info("（把 home.png 放到应用目录就能看到今天的小彩蛋啦～）")

    st.markdown(
        """
        <div style="text-align:center; padding:18px; border-radius:14px;
                    background: linear-gradient(135deg,#fde2e4,#fad2e1,#cddafd);
                    color:#5a3e6b; font-size:18px; line-height:1.7;">
        🌸 <b>今天也辛苦啦！</b> 🌸<br/>
        两份文件都安全到家了 ✨<br/>
        合上电脑，给自己一个大大的拥抱 🤗<br/>
        热茶、晚饭、还有想见的人，都在等你 🍵🍰💖<br/>
        <i>明天又是元气满满的一天～</i>
        </div>
        """,
        unsafe_allow_html=True,
    )


# Initialize session-state flags up front so reads never KeyError.
for key in ("excel_downloaded", "image_downloaded", "celebrated"):
    st.session_state.setdefault(key, False)


st.title("Excel 汇总工具")
st.caption(
    "上传 Excel 文件后，自动读取 Sheet1 的 G:P 区域，按固定名单汇总，生成新的结果 Excel，并导出高清表格图片。"
)

with st.expander("固定名单", expanded=False):
    st.write("、".join(FIXED_LIST))

uploaded_file = st.file_uploader(
    "上传 Excel 文件",
    type=["xlsx", "xlsm"],
    accept_multiple_files=False,
    help="要求工作表名为 Sheet1，G列为姓名，H:P 为 9 个数值指标。",
)

sheet_name = st.text_input("工作表名称", value=DEFAULT_SHEET_NAME)
custom_excel_filename = st.text_input(
    "输出 Excel 文件名",
    value=DEFAULT_EXCEL_FILENAME,
    help="可自定义输出文件名。默认格式为 业绩表-YYYYMMDD.xlsx。",
)

if uploaded_file is not None:
    st.info(f"已上传文件：{uploaded_file.name}")

    if st.button("开始处理", type="primary"):
        reset_download_flags()
        try:
            excel_bytes, result_df = process_excel(uploaded_file, sheet_name=sheet_name)
            excel_filename = sanitize_filename(custom_excel_filename, ".xlsx")
            plot_filename = build_plot_filename(excel_filename)
            plot_bytes = render_table_plot(result_df)

            # Stash in session state so the download buttons survive reruns.
            st.session_state.excel_bytes = excel_bytes
            st.session_state.plot_bytes = plot_bytes
            st.session_state.excel_filename = excel_filename
            st.session_state.plot_filename = plot_filename
            st.session_state.result_df = result_df
            st.session_state.has_result = True
        except Exception as exc:
            st.session_state.has_result = False
            st.error(str(exc))

    if st.session_state.get("has_result", False):
        st.success("处理完成。你可以预览结果，并分别下载 Excel 与高清图片。")
        st.dataframe(st.session_state.result_df, use_container_width=True)
        st.image(
            st.session_state.plot_bytes,
            caption="结果表格高清图片预览",
            use_container_width=True,
        )

        col1, col2 = st.columns(2)
        with col1:
            if st.download_button(
                label="下载结果 Excel",
                data=st.session_state.excel_bytes,
                file_name=st.session_state.excel_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="dl_excel",
            ):
                st.session_state.excel_downloaded = True
        with col2:
            if st.download_button(
                label="下载高清图片",
                data=st.session_state.plot_bytes,
                file_name=st.session_state.plot_filename,
                mime="image/png",
                use_container_width=True,
                key="dl_image",
            ):
                st.session_state.image_downloaded = True

        if CHINESE_FONT_PROP is None:
            st.warning(
                "当前运行环境中未检测到明确的中文字体。若图片中的中文显示异常，请把 NotoSansCJKsc-Regular.otf、"
                "SourceHanSansCN-Regular.otf 或 SimHei.ttf 放到应用目录或 fonts 目录中后重新部署。"
            )

        if st.session_state.excel_downloaded and st.session_state.image_downloaded:
            show_celebration()
else:
    st.write("请先上传一个 Excel 文件。")
